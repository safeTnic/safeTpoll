import json
import os
import mimetypes
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_protect
from django.db import transaction
from django.utils import timezone
from django.conf import settings

from .models import (
    Poll, Block, TextBlock, Question, QuestionOption,
    MatrixRow, MatrixColumn, PollResponse, Answer, UserTeamsSettings,
)
from .forms import UserTeamsSettingsForm


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def can_create_polls(user):
    return user.is_superuser or user.groups.filter(name='Poll Creators').exists()

def get_blocks_json(poll):
    """Return all blocks of a poll as a serializable list."""
    blocks = []
    for block in poll.blocks.all():
        b = {
            'id': block.id,
            'type': block.block_type,
            'page': block.page,
            'order': block.order,
        }
        if block.block_type == 'text':
            try:
                b['content'] = block.text_content.content
            except TextBlock.DoesNotExist:
                b['content'] = ''
        elif block.block_type == 'question':
            try:
                q = block.question
                b['question'] = {
                    'id': q.id,
                    'title': q.title,
                    'description': q.description,
                    'question_type': q.question_type,
                    'required': q.required,
                    'settings': q.settings,
                    'options': [
                        {'id': o.id, 'text': o.text, 'order': o.order}
                        for o in q.options.all()
                    ],
                    'matrix_rows': [
                        {'id': r.id, 'text': r.text, 'order': r.order}
                        for r in q.matrix_rows.all()
                    ],
                    'matrix_columns': [
                        {'id': c.id, 'text': c.text, 'order': c.order}
                        for c in q.matrix_columns.all()
                    ],
                }
            except Question.DoesNotExist:
                pass
        blocks.append(b)
    return blocks


def get_results_data(poll):
    """Compute statistics for all questions in a poll."""
    total_responses = poll.responses.filter(completed=True).count()
    question_stats = []

    for block in poll.blocks.filter(block_type='question'):
        try:
            q = block.question
        except Question.DoesNotExist:
            continue

        answers_qs = Answer.objects.filter(
            question=q,
            response__poll=poll,
            response__completed=True,
        )
        answer_count = answers_qs.count()

        stat = {
            'question_id': q.id,
            'title': q.title,
            'question_type': q.question_type,
            'answer_count': answer_count,
        }

        if q.question_type in ('single_choice', 'multiple_choice'):
            options = list(q.options.values_list('id', 'text'))
            counts = {opt_id: 0 for opt_id, _ in options}
            for answer in answers_qs:
                val = answer.value
                if isinstance(val, list):
                    for v in val:
                        try:
                            key = int(v)
                            if key in counts:
                                counts[key] += 1
                        except (ValueError, TypeError):
                            pass
                elif val is not None:
                    try:
                        key = int(val)
                        if key in counts:
                            counts[key] += 1
                    except (ValueError, TypeError):
                        pass
            stat['options'] = [
                {'id': opt_id, 'text': text, 'count': counts.get(opt_id, 0)}
                for opt_id, text in options
            ]

        elif q.question_type in ('text', 'textarea', 'date'):
            responses_list = []
            for answer in answers_qs[:100]:
                if answer.value:
                    responses_list.append(str(answer.value))
            stat['responses'] = responses_list

        elif q.question_type == 'rating':
            settings_data = q.settings or {}
            min_val = settings_data.get('min', 1)
            max_val = settings_data.get('max', 5)
            distribution = {i: 0 for i in range(min_val, max_val + 1)}
            total = 0
            count = 0
            for answer in answers_qs:
                try:
                    v = float(answer.value)
                    key = int(v)
                    if key in distribution:
                        distribution[key] += 1
                    total += v
                    count += 1
                except (ValueError, TypeError):
                    pass
            stat['average'] = round(total / count, 2) if count else None
            stat['distribution'] = [
                {'value': k, 'count': v} for k, v in sorted(distribution.items())
            ]

        elif q.question_type == 'number':
            values = []
            for answer in answers_qs:
                try:
                    values.append(float(answer.value))
                except (ValueError, TypeError):
                    pass
            stat['average'] = round(sum(values) / len(values), 2) if values else None
            stat['values'] = values[:200]

        elif q.question_type == 'ranking':
            options = list(q.options.values_list('id', 'text'))
            rank_sums = {opt_id: 0 for opt_id, _ in options}
            rank_counts = {opt_id: 0 for opt_id, _ in options}
            for answer in answers_qs:
                val = answer.value
                if isinstance(val, list):
                    for pos, opt_id in enumerate(val):
                        try:
                            key = int(opt_id)
                            if key in rank_sums:
                                rank_sums[key] += pos + 1
                                rank_counts[key] += 1
                        except (ValueError, TypeError):
                            pass
            stat['options'] = [
                {
                    'id': opt_id,
                    'text': text,
                    'average_rank': round(rank_sums[opt_id] / rank_counts[opt_id], 2)
                    if rank_counts[opt_id] else None,
                }
                for opt_id, text in options
            ]

        elif q.question_type == 'file':
            files = []
            for answer in answers_qs:
                if answer.file:
                    raw_name = os.path.basename(answer.file.name)
                    ext = raw_name.rsplit('.', 1)[-1].lower() if '.' in raw_name else ''
                    if ext == 'pdf':
                        file_type = 'pdf'
                    elif ext in ('xlsx', 'xls', 'csv'):
                        file_type = 'excel'
                    elif ext in ('docx', 'doc', 'odt', 'rtf'):
                        file_type = 'word'
                    elif ext in ('pptx', 'ppt'):
                        file_type = 'ppt'
                    elif ext in ('png', 'jpg', 'jpeg', 'gif', 'webp', 'svg'):
                        file_type = 'image'
                    elif ext in ('zip', 'rar', '7z', 'tar', 'gz'):
                        file_type = 'archive'
                    else:
                        file_type = 'generic'
                    files.append({
                        'url': answer.file.url,
                        'name': raw_name,
                        'ext': ext.upper() or 'FILE',
                        'file_type': file_type,
                    })
            stat['files'] = files

        elif q.question_type == 'matrix':
            rows = list(q.matrix_rows.values_list('id', 'text'))
            cols = list(q.matrix_columns.values_list('id', 'text'))
            matrix = {r_id: {c_id: 0 for c_id, _ in cols} for r_id, _ in rows}
            for answer in answers_qs:
                val = answer.value
                if isinstance(val, dict):
                    for row_id_str, col_id in val.items():
                        try:
                            row_id = int(row_id_str)
                            col_id_int = int(col_id)
                            if row_id in matrix and col_id_int in matrix[row_id]:
                                matrix[row_id][col_id_int] += 1
                        except (ValueError, TypeError):
                            pass
            stat['rows'] = [{'id': r_id, 'text': r_text} for r_id, r_text in rows]
            stat['columns'] = [{'id': c_id, 'text': c_text} for c_id, c_text in cols]
            stat['matrix'] = {
                str(r_id): {str(c_id): cnt for c_id, cnt in col_data.items()}
                for r_id, col_data in matrix.items()
            }

        question_stats.append(stat)

    return {
        'total_responses': total_responses,
        'questions': question_stats,
    }


def _get_poll_for_participant(poll_id):
    """Look up a poll by custom_id or guid."""
    try:
        return Poll.objects.get(custom_id=poll_id)
    except Poll.DoesNotExist:
        pass
    try:
        return Poll.objects.get(guid=poll_id)
    except (Poll.DoesNotExist, ValueError):
        raise Http404("Umfrage nicht gefunden.")


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@login_required
def dashboard(request):
    polls = Poll.objects.filter(creator=request.user).prefetch_related('responses')
    poll_data = []
    for poll in polls:
        poll_data.append({
            'poll': poll,
            'response_count': poll.total_responses(),
        })
    return render(request, 'polls/dashboard.html', {
        'poll_data': poll_data,
        'user_can_create_polls': can_create_polls(request.user),
    })


@login_required
def pinnwand(request):
    pinned_polls = (
        Poll.objects
        .filter(pinned=True)
        .select_related('creator')
        .prefetch_related('responses')
        .order_by('-updated_at')
    )
    poll_data = [
        {'poll': p, 'response_count': p.total_responses()}
        for p in pinned_polls
    ]
    return render(request, 'polls/pinnwand.html', {'poll_data': poll_data})


@login_required
@require_POST
def poll_pin_toggle(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    poll.pinned = not poll.pinned
    poll.save(update_fields=['pinned'])
    if 'application/json' in request.META.get('HTTP_ACCEPT', ''):
        return JsonResponse({'status': 'ok', 'pinned': poll.pinned})
    return redirect('dashboard')


# ---------------------------------------------------------------------------
# Poll CRUD
# ---------------------------------------------------------------------------

@login_required
@require_POST
def poll_create(request):
    if not can_create_polls(request.user):
        messages.error(request, 'Sie haben keine Berechtigung, Umfragen zu erstellen.')
        return redirect('dashboard')
    title = request.POST.get('title', '').strip()
    if not title:
        messages.error(request, 'Bitte geben Sie einen Titel ein.')
        return redirect('dashboard')
    poll = Poll.objects.create(title=title, creator=request.user)
    poll.custom_id = str(poll.guid)
    poll.save(update_fields=['custom_id'])
    return redirect('poll_editor', poll_id=poll.id)


@login_required
def poll_editor(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    blocks_json = json.dumps(get_blocks_json(poll))
    poll_url = request.build_absolute_uri(f'/tool/{poll.get_url_id()}/')
    return render(request, 'polls/editor.html', {
        'poll': poll,
        'blocks_json': blocks_json,
        'poll_url': poll_url,
    })


@login_required
@require_POST
@csrf_protect
def poll_settings_update(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Ungültige JSON-Daten'}, status=400)

    if 'title' in data:
        poll.title = str(data['title'])[:255]
    if 'status' in data and data['status'] in ('draft', 'active', 'closed'):
        poll.status = data['status']
    if 'is_anonymous' in data:
        poll.is_anonymous = bool(data['is_anonymous'])
    if 'show_vote_count' in data:
        poll.show_vote_count = bool(data['show_vote_count'])
    if 'allow_multiple_responses' in data:
        poll.allow_multiple_responses = bool(data['allow_multiple_responses'])
    if 'questions_per_page' in data:
        try:
            poll.questions_per_page = max(0, int(data['questions_per_page']))
        except (ValueError, TypeError):
            pass
    if 'custom_id' in data:
        val = str(data['custom_id']).strip()[:100] or str(poll.guid)
        if Poll.objects.filter(custom_id=val).exclude(id=poll.id).exists():
            return JsonResponse({'error': 'Diese ID wird bereits verwendet.'}, status=400)
        poll.custom_id = val
    if 'start_date' in data:
        try:
            poll.start_date = datetime.fromisoformat(data['start_date']) if data['start_date'] else None
        except (ValueError, TypeError):
            poll.start_date = None
    if 'end_date' in data:
        try:
            poll.end_date = datetime.fromisoformat(data['end_date']) if data['end_date'] else None
        except (ValueError, TypeError):
            poll.end_date = None

    poll.save()
    return JsonResponse({'status': 'ok', 'poll': {
        'id': poll.id,
        'title': poll.title,
        'status': poll.status,
        'url_id': poll.get_url_id(),
    }})


@login_required
@require_POST
@csrf_protect
def save_blocks(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    try:
        data = json.loads(request.body)
        incoming_blocks = data.get('blocks', [])
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Ungültige JSON-Daten'}, status=400)

    with transaction.atomic():
        # Track which existing block IDs are being kept
        kept_block_ids = set()
        for b_data in incoming_blocks:
            b_id = b_data.get('id')
            if isinstance(b_id, int):
                kept_block_ids.add(b_id)

        # Delete blocks not in incoming list — only if they have no answers
        for existing_block in poll.blocks.all():
            if existing_block.id not in kept_block_ids:
                has_answers = False
                if existing_block.block_type == 'question':
                    try:
                        has_answers = existing_block.question.answers.exists()
                    except Question.DoesNotExist:
                        pass
                if not has_answers:
                    existing_block.delete()

        # Create / update blocks
        result_blocks = []
        for order_idx, b_data in enumerate(incoming_blocks):
            b_id = b_data.get('id')
            b_type = b_data.get('type', 'text')
            b_page = int(b_data.get('page', 1))
            b_order = int(b_data.get('order', order_idx))

            if isinstance(b_id, int):
                try:
                    block = Block.objects.get(id=b_id, poll=poll)
                    block.order = b_order
                    block.page = b_page
                    block.save()
                except Block.DoesNotExist:
                    block = Block.objects.create(
                        poll=poll, block_type=b_type, order=b_order, page=b_page
                    )
            else:
                block = Block.objects.create(
                    poll=poll, block_type=b_type, order=b_order, page=b_page
                )

            if b_type == 'text':
                content = b_data.get('content', '')
                TextBlock.objects.update_or_create(
                    block=block,
                    defaults={'content': content},
                )

            elif b_type == 'question':
                q_data = b_data.get('question', {})
                q_id = q_data.get('id')
                q_type = q_data.get('question_type', 'single_choice')
                q_title = str(q_data.get('title', ''))[:500]
                q_desc = str(q_data.get('description', ''))
                q_required = bool(q_data.get('required', True))
                q_settings = q_data.get('settings', {})
                if not isinstance(q_settings, dict):
                    q_settings = {}

                if q_id:
                    try:
                        question = Question.objects.get(id=q_id, block=block)
                        question.title = q_title
                        question.description = q_desc
                        question.question_type = q_type
                        question.required = q_required
                        question.settings = q_settings
                        question.save()
                    except Question.DoesNotExist:
                        question = Question.objects.create(
                            block=block, title=q_title, description=q_desc,
                            question_type=q_type, required=q_required, settings=q_settings,
                        )
                else:
                    question, _ = Question.objects.get_or_create(
                        block=block,
                        defaults={
                            'title': q_title, 'description': q_desc,
                            'question_type': q_type, 'required': q_required,
                            'settings': q_settings,
                        }
                    )
                    question.title = q_title
                    question.description = q_desc
                    question.question_type = q_type
                    question.required = q_required
                    question.settings = q_settings
                    question.save()

                # Update options
                incoming_options = q_data.get('options', [])
                kept_opt_ids = {o['id'] for o in incoming_options if o.get('id')}
                question.options.exclude(id__in=kept_opt_ids).delete()
                for o_data in incoming_options:
                    o_id = o_data.get('id')
                    o_text = str(o_data.get('text', ''))[:500]
                    o_order = int(o_data.get('order', 0))
                    if o_id:
                        QuestionOption.objects.filter(id=o_id, question=question).update(
                            text=o_text, order=o_order
                        )
                    else:
                        QuestionOption.objects.create(
                            question=question, text=o_text, order=o_order
                        )

                # Update matrix rows
                incoming_rows = q_data.get('matrix_rows', [])
                kept_row_ids = {r['id'] for r in incoming_rows if r.get('id')}
                question.matrix_rows.exclude(id__in=kept_row_ids).delete()
                for r_data in incoming_rows:
                    r_id = r_data.get('id')
                    r_text = str(r_data.get('text', ''))[:500]
                    r_order = int(r_data.get('order', 0))
                    if r_id:
                        MatrixRow.objects.filter(id=r_id, question=question).update(
                            text=r_text, order=r_order
                        )
                    else:
                        MatrixRow.objects.create(question=question, text=r_text, order=r_order)

                # Update matrix columns
                incoming_cols = q_data.get('matrix_columns', [])
                kept_col_ids = {c['id'] for c in incoming_cols if c.get('id')}
                question.matrix_columns.exclude(id__in=kept_col_ids).delete()
                for c_data in incoming_cols:
                    c_id = c_data.get('id')
                    c_text = str(c_data.get('text', ''))[:500]
                    c_order = int(c_data.get('order', 0))
                    if c_id:
                        MatrixColumn.objects.filter(id=c_id, question=question).update(
                            text=c_text, order=c_order
                        )
                    else:
                        MatrixColumn.objects.create(question=question, text=c_text, order=c_order)

            result_blocks.append(block.id)

    poll.refresh_from_db()
    return JsonResponse({'status': 'ok', 'blocks': get_blocks_json(poll)})


@login_required
@require_POST
def poll_publish(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    poll.status = 'active'
    poll.save()
    if 'application/json' in request.META.get('HTTP_ACCEPT', ''):
        return JsonResponse({'status': 'ok'})
    return redirect('dashboard')


@login_required
@require_POST
def poll_close(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    poll.status = 'closed'
    poll.save()
    if 'application/json' in request.META.get('HTTP_ACCEPT', ''):
        return JsonResponse({'status': 'ok'})
    return redirect('dashboard')


@login_required
@require_POST
def poll_delete(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    poll.delete()
    messages.success(request, 'Umfrage wurde gelöscht.')
    return redirect('dashboard')


# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------

@login_required
def poll_results(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    results = get_results_data(poll)
    teams_configured = False
    try:
        teams_configured = request.user.teams_settings.is_configured()
    except UserTeamsSettings.DoesNotExist:
        pass
    responses = (
        PollResponse.objects
        .filter(poll=poll, completed=True)
        .select_related('user')
        .order_by('-submitted_at')
    )
    return render(request, 'polls/results.html', {
        'poll': poll,
        'results': results,
        'teams_configured': teams_configured,
        'voter_responses': responses,
    })


@login_required
def response_detail(request, poll_id, response_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    response_obj = get_object_or_404(PollResponse, id=response_id, poll=poll, completed=True)

    all_ids = list(
        PollResponse.objects.filter(poll=poll, completed=True)
        .order_by('-submitted_at')
        .values_list('id', flat=True)
    )
    try:
        idx = all_ids.index(response_obj.id)
    except ValueError:
        idx = 0

    prev_id = all_ids[idx - 1] if idx > 0 else None
    next_id = all_ids[idx + 1] if idx < len(all_ids) - 1 else None

    answer_map = {a.question_id: a for a in response_obj.answers.select_related('question').all()}

    answer_rows = []
    for block in poll.blocks.filter(block_type='question').order_by('order'):
        try:
            q = block.question
        except Question.DoesNotExist:
            continue

        answer = answer_map.get(q.id)
        formatted = None

        if answer:
            val = answer.value
            if q.question_type == 'single_choice':
                try:
                    opt = q.options.get(id=int(val))
                    formatted = opt.text
                except Exception:
                    formatted = str(val) if val is not None else None

            elif q.question_type == 'multiple_choice':
                try:
                    ids = [int(v) for v in (val if isinstance(val, list) else [val])]
                    opts = {o.id: o.text for o in q.options.filter(id__in=ids)}
                    formatted = [opts.get(i, str(i)) for i in ids]
                except Exception:
                    formatted = val

            elif q.question_type == 'ranking':
                try:
                    ids = [int(v) for v in val]
                    opts = {o.id: o.text for o in q.options.all()}
                    formatted = [opts.get(i, str(i)) for i in ids]
                except Exception:
                    formatted = val

            elif q.question_type == 'matrix':
                rows = {str(r.id): r.text for r in q.matrix_rows.all()}
                cols = {str(c.id): c.text for c in q.matrix_columns.all()}
                if isinstance(val, dict):
                    formatted = {rows.get(k, k): cols.get(str(v2), str(v2)) for k, v2 in val.items()}
                else:
                    formatted = val

            elif q.question_type == 'file':
                formatted = {'url': answer.file.url, 'name': os.path.basename(answer.file.name)} if answer.file else None

            elif q.question_type == 'rating':
                formatted = val

            else:
                formatted = str(val) if val is not None else None

        answer_rows.append({
            'question': q,
            'answer': answer,
            'formatted': formatted,
        })

    return render(request, 'polls/response_detail.html', {
        'poll': poll,
        'response': response_obj,
        'answer_rows': answer_rows,
        'prev_id': prev_id,
        'next_id': next_id,
        'current_num': idx + 1,
        'total': len(all_ids),
    })


@login_required
@require_POST
def delete_response(request, poll_id, response_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    resp = get_object_or_404(PollResponse, id=response_id, poll=poll, completed=True)
    resp.delete()
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f'poll_results_{poll.id}',
                {'type': 'poll_results_update', 'data': get_results_data(poll)},
            )
    except Exception:
        pass
    return JsonResponse({'status': 'ok'})


@login_required
def poll_results_data(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    return JsonResponse(get_results_data(poll))


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@login_required
def poll_export_excel(request, poll_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = 'Zusammenfassung'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

    ws_summary['A1'] = 'Umfrage'
    ws_summary['B1'] = poll.title
    ws_summary['A2'] = 'Status'
    ws_summary['B2'] = poll.get_status_display()
    ws_summary['A3'] = 'Erstellt am'
    ws_summary['B3'] = poll.created_at.strftime('%d.%m.%Y %H:%M')
    ws_summary['A4'] = 'Gesamtantworten'
    ws_summary['B4'] = poll.total_responses()
    ws_summary['A5'] = 'Anonym'
    ws_summary['B5'] = 'Ja' if poll.is_anonymous else 'Nein'
    ws_summary['A6'] = 'Ersteller'
    ws_summary['B6'] = poll.creator.get_full_name() or poll.creator.username

    for row in ws_summary.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40

    # Sheets per question
    completed_responses = PollResponse.objects.filter(
        poll=poll, completed=True
    ).prefetch_related('answers__question')

    for block in poll.blocks.filter(block_type='question'):
        try:
            q = block.question
        except Question.DoesNotExist:
            continue

        # Sanitize: Excel forbids \ / : * ? [ ] in sheet names
        import re
        safe = re.sub(r'[\\/:*?\[\]]', '', q.title).strip() or 'Frage'
        sheet_title = (safe[:28] + '..') if len(safe) > 30 else safe
        # Ensure uniqueness and max 31 chars
        existing_titles = [s.title for s in wb.worksheets]
        base_title = sheet_title
        counter = 1
        while sheet_title in existing_titles:
            sheet_title = f"{base_title[:27]}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=sheet_title)
        ws['A1'] = q.title
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = q.get_question_type_display()
        ws['A2'].font = Font(italic=True, color='666666')

        if q.question_type in ('single_choice', 'multiple_choice'):
            options = list(q.options.all())
            ws['A4'] = 'Option'
            ws['B4'] = 'Anzahl'
            ws['A4'].font = header_font
            ws['A4'].fill = header_fill
            ws['B4'].font = header_font
            ws['B4'].fill = header_fill
            counts = {o.id: 0 for o in options}
            for resp in completed_responses:
                ans = resp.answers.filter(question=q).first()
                if ans and ans.value is not None:
                    val = ans.value
                    if isinstance(val, list):
                        for v in val:
                            try:
                                counts[int(v)] = counts.get(int(v), 0) + 1
                            except (ValueError, TypeError):
                                pass
                    else:
                        try:
                            counts[int(val)] = counts.get(int(val), 0) + 1
                        except (ValueError, TypeError):
                            pass
            for row_idx, opt in enumerate(options, start=5):
                ws.cell(row=row_idx, column=1, value=opt.text)
                ws.cell(row=row_idx, column=2, value=counts.get(opt.id, 0))

        elif q.question_type in ('text', 'textarea', 'number', 'date'):
            ws['A4'] = 'Antwort'
            ws['A4'].font = header_font
            ws['A4'].fill = header_fill
            row_idx = 5
            for resp in completed_responses:
                ans = resp.answers.filter(question=q).first()
                if ans and ans.value is not None:
                    ws.cell(row=row_idx, column=1, value=str(ans.value))
                    row_idx += 1

        elif q.question_type == 'rating':
            ws['A4'] = 'Bewertung'
            ws['B4'] = 'Anzahl'
            ws['A4'].font = header_font
            ws['A4'].fill = header_fill
            ws['B4'].font = header_font
            ws['B4'].fill = header_fill
            q_settings = q.settings or {}
            min_val = q_settings.get('min', 1)
            max_val = q_settings.get('max', 5)
            dist = {i: 0 for i in range(min_val, max_val + 1)}
            for resp in completed_responses:
                ans = resp.answers.filter(question=q).first()
                if ans and ans.value is not None:
                    try:
                        v = int(float(ans.value))
                        if v in dist:
                            dist[v] += 1
                    except (ValueError, TypeError):
                        pass
            for row_idx, (val, cnt) in enumerate(sorted(dist.items()), start=5):
                ws.cell(row=row_idx, column=1, value=val)
                ws.cell(row=row_idx, column=2, value=cnt)

        elif q.question_type == 'ranking':
            options = list(q.options.all())
            ws['A4'] = 'Option'
            ws['B4'] = 'Ø Rang'
            ws['A4'].font = header_font
            ws['A4'].fill = header_fill
            ws['B4'].font = header_font
            ws['B4'].fill = header_fill
            rank_sums = {o.id: 0 for o in options}
            rank_counts = {o.id: 0 for o in options}
            for resp in completed_responses:
                ans = resp.answers.filter(question=q).first()
                if ans and isinstance(ans.value, list):
                    for pos, opt_id in enumerate(ans.value):
                        try:
                            key = int(opt_id)
                            if key in rank_sums:
                                rank_sums[key] += pos + 1
                                rank_counts[key] += 1
                        except (ValueError, TypeError):
                            pass
            for row_idx, opt in enumerate(options, start=5):
                ws.cell(row=row_idx, column=1, value=opt.text)
                avg = round(rank_sums[opt.id] / rank_counts[opt.id], 2) if rank_counts[opt.id] else None
                ws.cell(row=row_idx, column=2, value=avg)

        elif q.question_type == 'matrix':
            rows = list(q.matrix_rows.all())
            cols = list(q.matrix_columns.all())
            ws.cell(row=4, column=1, value='Zeile / Spalte')
            ws.cell(row=4, column=1).font = header_font
            ws.cell(row=4, column=1).fill = header_fill
            for c_idx, col in enumerate(cols, start=2):
                ws.cell(row=4, column=c_idx, value=col.text)
                ws.cell(row=4, column=c_idx).font = header_font
                ws.cell(row=4, column=c_idx).fill = header_fill
            matrix = {r.id: {c.id: 0 for c in cols} for r in rows}
            for resp in completed_responses:
                ans = resp.answers.filter(question=q).first()
                if ans and isinstance(ans.value, dict):
                    for r_id_str, c_id in ans.value.items():
                        try:
                            r_id = int(r_id_str)
                            c_id_int = int(c_id)
                            if r_id in matrix and c_id_int in matrix[r_id]:
                                matrix[r_id][c_id_int] += 1
                        except (ValueError, TypeError):
                            pass
            for r_idx, row in enumerate(rows, start=5):
                ws.cell(row=r_idx, column=1, value=row.text)
                for c_idx, col in enumerate(cols, start=2):
                    ws.cell(row=r_idx, column=c_idx, value=matrix[row.id].get(col.id, 0))

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_title = ''.join(c for c in poll.title if c.isalnum() or c in ' _-')[:50]
    response['Content-Disposition'] = f'attachment; filename="{safe_title}_Ergebnisse.xlsx"'
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# Teams export
# ---------------------------------------------------------------------------

@login_required
@require_POST
@csrf_protect
def teams_export(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)

    try:
        teams_settings = request.user.teams_settings
    except UserTeamsSettings.DoesNotExist:
        return JsonResponse({'error': 'Teams-Integration nicht konfiguriert.'}, status=400)

    if not teams_settings.is_configured():
        return JsonResponse({'error': 'Teams-Integration nicht vollständig konfiguriert.'}, status=400)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'Ungültige JSON-Daten'}, status=400)

    event_title = data.get('event_title', poll.title)
    start_datetime = data.get('start_datetime')
    end_datetime = data.get('end_datetime')
    description = data.get('description', '')

    if not start_datetime or not end_datetime:
        return JsonResponse({'error': 'Start- und Enddatum sind erforderlich.'}, status=400)

    try:
        import msal
        import urllib.request
        import urllib.error

        authority = f"https://login.microsoftonline.com/{teams_settings.tenant_id}"
        app = msal.ConfidentialClientApplication(
            teams_settings.client_id,
            authority=authority,
            client_credential=teams_settings.client_secret,
        )

        result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])

        if 'access_token' not in result:
            error_desc = result.get('error_description', 'Unbekannter Fehler')
            return JsonResponse({'error': f'Token-Fehler: {error_desc}'}, status=400)

        token = result['access_token']

        event_body = {
            "subject": event_title,
            "body": {
                "contentType": "HTML",
                "content": description or f"Umfrage: {poll.title}",
            },
            "start": {
                "dateTime": start_datetime,
                "timeZone": "Europe/Berlin",
            },
            "end": {
                "dateTime": end_datetime,
                "timeZone": "Europe/Berlin",
            },
            "isOnlineMeeting": True,
            "onlineMeetingProvider": "teamsForBusiness",
        }

        req_data = json.dumps(event_body).encode('utf-8')
        req = urllib.request.Request(
            "https://graph.microsoft.com/v1.0/me/events",
            data=req_data,
            method='POST',
            headers={
                'Authorization': f'Bearer {token}',
                'Content-Type': 'application/json',
            }
        )

        try:
            with urllib.request.urlopen(req) as resp:
                resp_data = json.loads(resp.read().decode('utf-8'))
            return JsonResponse({
                'status': 'ok',
                'event_id': resp_data.get('id'),
                'join_url': resp_data.get('onlineMeeting', {}).get('joinUrl', ''),
            })
        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8')
            return JsonResponse({'error': f'Graph API Fehler: {error_body}'}, status=400)

    except Exception as e:
        return JsonResponse({'error': f'Fehler: {str(e)}'}, status=500)


# ---------------------------------------------------------------------------
# Poll participant view
# ---------------------------------------------------------------------------

@login_required
def poll_view(request, poll_id):
    poll = _get_poll_for_participant(poll_id)

    if not poll.is_active():
        status_messages = {
            'draft': 'Diese Umfrage ist noch nicht veröffentlicht.',
            'closed': 'Diese Umfrage ist bereits geschlossen.',
            'active': 'Diese Umfrage ist derzeit nicht aktiv (Zeitraum beachten).',
        }
        return render(request, 'polls/poll_view.html', {
            'poll': poll,
            'not_active': True,
            'not_active_message': status_messages.get(poll.status, 'Diese Umfrage ist nicht verfügbar.'),
        })

    # Check if user already responded
    if not poll.allow_multiple_responses:
        existing = PollResponse.objects.filter(
            poll=poll, user=request.user, completed=True
        ).first()
        if existing:
            return render(request, 'polls/poll_view.html', {
                'poll': poll,
                'already_responded': True,
            })

    # Get or create in-progress response
    response_obj, _ = PollResponse.objects.get_or_create(
        poll=poll,
        user=request.user,
        completed=False,
    )

    current_page = int(request.GET.get('page', response_obj.current_page))
    blocks = poll.blocks.all()

    # Separate text blocks and question blocks
    all_question_blocks = [b for b in blocks if b.block_type == 'question']
    all_text_blocks = [b for b in blocks if b.block_type == 'text']

    qpp = poll.questions_per_page
    if qpp > 0:
        total_pages = max(1, -(-len(all_question_blocks) // qpp))  # ceiling division
        start_idx = (current_page - 1) * qpp
        end_idx = start_idx + qpp
        page_question_blocks = all_question_blocks[start_idx:end_idx]
    else:
        total_pages = 1
        page_question_blocks = all_question_blocks

    # Build display blocks for current page (text + questions)
    display_blocks = list(all_text_blocks) + list(page_question_blocks)
    display_blocks.sort(key=lambda b: b.order)

    # Get existing answers for this response
    existing_answers = {
        ans.question_id: ans.value
        for ans in response_obj.answers.all()
    }

    response_obj.current_page = current_page
    response_obj.save()

    return render(request, 'polls/poll_view.html', {
        'poll': poll,
        'response_obj': response_obj,
        'display_blocks': display_blocks,
        'current_page': current_page,
        'total_pages': total_pages,
        'existing_answers': json.dumps(existing_answers),
        'is_last_page': current_page >= total_pages,
    })


@login_required
@require_POST
@csrf_protect
def poll_submit(request, poll_id):
    poll = _get_poll_for_participant(poll_id)

    if not poll.is_active():
        return JsonResponse({'error': 'Umfrage ist nicht aktiv.'}, status=400)

    content_type = request.content_type or ''
    files_data = {}

    if 'application/json' in content_type:
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({'error': 'Ungültige Daten.'}, status=400)
        current_page = int(data.get('current_page', 1))
        answers_data = data.get('answers', {})
    else:
        # multipart/form-data (used when file uploads are present)
        try:
            current_page = int(request.POST.get('current_page', 1))
        except (ValueError, TypeError):
            current_page = 1
        answers_data = {}
        for key, value in request.POST.items():
            if key.startswith('answer_'):
                q_id = key[7:]
                try:
                    answers_data[q_id] = json.loads(value)
                except (json.JSONDecodeError, ValueError):
                    answers_data[q_id] = value
        for key, uploaded_file in request.FILES.items():
            if key.startswith('file_'):
                q_id = key[5:]
                files_data[q_id] = uploaded_file

    # Get or create response
    response_obj, _ = PollResponse.objects.get_or_create(
        poll=poll,
        user=request.user,
        completed=False,
    )

    # Save answers for current page questions
    all_question_blocks = list(poll.blocks.filter(block_type='question'))
    qpp = poll.questions_per_page
    if qpp > 0:
        start_idx = (current_page - 1) * qpp
        end_idx = start_idx + qpp
        page_question_blocks = all_question_blocks[start_idx:end_idx]
        total_pages = max(1, -(-len(all_question_blocks) // qpp))
    else:
        page_question_blocks = all_question_blocks
        total_pages = 1

    for block in page_question_blocks:
        try:
            question = block.question
        except Question.DoesNotExist:
            continue

        q_id_str = str(question.id)
        if q_id_str in answers_data:
            value = answers_data[q_id_str]
            Answer.objects.update_or_create(
                response=response_obj,
                question=question,
                defaults={'value': value},
            )
        if q_id_str in files_data:
            uploaded_file = files_data[q_id_str]
            # Validate file size
            max_size = getattr(settings, 'MAX_UPLOAD_SIZE', 10 * 1024 * 1024)
            if uploaded_file.size <= max_size:
                ans, _ = Answer.objects.get_or_create(
                    response=response_obj,
                    question=question,
                    defaults={'value': uploaded_file.name},
                )
                ans.file = uploaded_file
                ans.value = uploaded_file.name
                ans.save()

    is_last_page = current_page >= total_pages

    if is_last_page:
        response_obj.completed = True
        response_obj.submitted_at = timezone.now()
        response_obj.current_page = current_page
        response_obj.save()

        # Broadcast via WebSocket
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            channel_layer = get_channel_layer()
            if channel_layer:
                async_to_sync(channel_layer.group_send)(
                    f'poll_results_{poll.id}',
                    {
                        'type': 'poll_results_update',
                        'data': get_results_data(poll),
                    }
                )
        except Exception:
            pass

        return JsonResponse({'completed': True, 'next_page': None})
    else:
        response_obj.current_page = current_page + 1
        response_obj.save()
        return JsonResponse({'completed': False, 'next_page': current_page + 1})


# ---------------------------------------------------------------------------
# Email reminder
# ---------------------------------------------------------------------------

@login_required
@require_http_methods(['GET', 'POST'])
def poll_remind(request, poll_id):
    poll = get_object_or_404(Poll, id=poll_id, creator=request.user)

    voted_user_ids = set(
        PollResponse.objects.filter(poll=poll, completed=True)
        .values_list('user_id', flat=True)
    )
    non_voters = (
        User.objects.filter(is_active=True)
        .exclude(id__in=voted_user_ids)
        .exclude(id=request.user.id)
        .exclude(email='')
        .order_by('last_name', 'first_name', 'username')
    )

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({'error': 'Ungültige Daten.'}, status=400)

        recipient_ids = data.get('recipient_ids')
        if recipient_ids:
            recipients = list(non_voters.filter(id__in=recipient_ids))
        else:
            recipients = list(non_voters)

        if not recipients:
            return JsonResponse({'error': 'Keine Empfänger ausgewählt.'}, status=400)

        poll_url = request.build_absolute_uri(f'/tool/{poll.get_url_id()}/')
        sent_count = 0
        failed_emails = []

        for user in recipients:
            try:
                send_mail(
                    subject=f'Erinnerung: Bitte nehmen Sie an der Umfrage teil – {poll.title}',
                    message=(
                        f'Hallo {user.get_full_name() or user.username},\n\n'
                        f'Sie haben noch nicht an der folgenden Umfrage teilgenommen:\n\n'
                        f'» {poll.title} «\n\n'
                        f'Bitte nehmen Sie hier teil:\n{poll_url}\n\n'
                        f'Mit freundlichen Grüßen\n'
                        f'{request.user.get_full_name() or request.user.username}'
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
                sent_count += 1
            except Exception:
                failed_emails.append(user.email)

        return JsonResponse({
            'status': 'ok',
            'sent': sent_count,
            'failed': failed_emails,
        })

    return JsonResponse({
        'non_voters': [
            {
                'id': u.id,
                'name': u.get_full_name() or u.username,
                'email': u.email,
            }
            for u in non_voters
        ]
    })


# ---------------------------------------------------------------------------
# User settings
# ---------------------------------------------------------------------------

@login_required
def user_settings(request):
    try:
        teams_settings = request.user.teams_settings
    except UserTeamsSettings.DoesNotExist:
        teams_settings = None

    if request.method == 'POST':
        form = UserTeamsSettingsForm(request.POST, instance=teams_settings)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.user = request.user
            # Preserve existing client_secret if field was left empty
            if not obj.client_secret and teams_settings and teams_settings.client_secret:
                obj.client_secret = teams_settings.client_secret
            obj.save()
            messages.success(request, 'Einstellungen gespeichert.')
            return redirect('user_settings')
    else:
        form = UserTeamsSettingsForm(instance=teams_settings)

    return render(request, 'polls/settings.html', {
        'form': form,
        'teams_settings': teams_settings,
    })
